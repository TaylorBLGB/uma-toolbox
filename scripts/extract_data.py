"""
Extracts public reference data from the source spreadsheets into JSON files
consumed by the static site. Deliberately drops personal collection-tracking
columns (Want?/Have?/Use?) - those are not published. "In Game?" is kept as
`inGame` since it's a public fact (whether the character/card is released
yet), not personal collection status.

Source files are NOT part of the repo (they contain personal collection data)
and their path differs per machine, so pass them explicitly:

  python scripts/extract_data.py "path/to/Datasheet.xlsx" "path/to/RL UMA.xlsx"

With no arguments it falls back to this machine's Downloads folder.
"""
import datetime
import json
import re
import sys
import openpyxl
from pathlib import Path

DEFAULT_DATASHEET = Path.home() / "Downloads" / "Datasheet.xlsx"
DEFAULT_RLUMA = Path.home() / "Downloads" / "RL UMA (1).xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"

DATASHEET = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_DATASHEET
RLUMA = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_RLUMA


def clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date().isoformat() if isinstance(v, datetime.datetime) else v.isoformat()
    if isinstance(v, datetime.time):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def extract_umas():
    wb = openpyxl.load_workbook(DATASHEET, data_only=True)
    ws = wb["Umas"]
    header = [ws.cell(row=2, column=c).value for c in range(1, 35)]
    idx = {h: i + 1 for i, h in enumerate(header) if h}

    out = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=idx["Name"]).value
        if not name:
            continue
        get = lambda col: clean(ws.cell(row=r, column=idx[col]).value) if col in idx else None
        out.append({
            "name": name,
            "inGame": get("In Game?") == "Y",
            "costume": get("Costume"),
            "version": get("Version"),
            "uniqueSkillName": get("Ult"),
            "uniqueCondition": get("Condition"),
            "preConditionNote": get("Pre-Condition/Note"),
            "distAptitude": get("Dist"),
            "style": get("Style"),
            "targetSpeedBonus": get("Target Speed"),
            "accelerationBonus": get("Acceleration"),
            "healBonus": get("Heal"),
            "uniqueEffectNote": get("Unique"),
            "duration": get("Duration"),
            "aptitude": {
                "turf": get("Turf"), "dirt": get("Dirt"),
                "sprint": get("Sprint"), "mile": get("Mile"),
                "medium": get("Med"), "long": get("Long"),
                "front": get("Front"), "pace": get("Pace"),
                "late": get("Late"), "end": get("End"),
            },
            "statBonus": {
                "spd": get("SPD"), "sta": get("STA"), "pwr": get("PWR"),
                "gut": get("GUT"), "wit": get("WIT"),
            },
            "note": get("Note"),
            "release": get("Release"),
        })
    wb.close()
    return out


def extract_supports():
    wb = openpyxl.load_workbook(DATASHEET, data_only=True)
    ws = wb["Supports"]
    header = [ws.cell(row=2, column=c).value for c in range(1, 27)]
    idx = {h: i + 1 for i, h in enumerate(header) if h}

    out = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=idx["Name"]).value
        if not name:
            continue
        get = lambda col: clean(ws.cell(row=r, column=idx[col]).value) if col in idx else None
        out.append({
            "character": get("Character"),
            "name": name,
            "inGame": get("In Game") == "Y",
            "type": get("Type"),
            "grade": get("Grade"),
            "lbPips": get("LB Pips"),
            "friendshipBonus": get("FBonus"),
            "motivationEffect": get("MEffect"),
            "trainingBonus": get("TBonus"),
            "effectMPlusTBonus": get("EffectM+TBonus"),
            "totalBonus": get("T+M+FBonus"),
            "initStat": get("Init Stat"),
            "initGauge": get("Init Gauge"),
            "raceBonus": get("RaceBonus"),
            "fanBonus": get("FanBonus"),
            "hintLevel": get("HintLvl"),
            "hintFrequency": get("Hint Freq"),
            "skillPtPriority": get("SPrio"),
            "skillPtBonus": get("SPBonus"),
            "notes": get("Notes"),
            "friendshipRatio": get("FRatio"),
            "release": get("Release"),
        })
    wb.close()
    return out


MONTHS = {
    "Jan": "January", "Feb": "February", "Mar": "March", "Apr": "April",
    "May": "May", "Jun": "June", "Jul": "July", "Aug": "August",
    "Sep": "September", "Oct": "October", "Nov": "November", "Dec": "December",
}


def slot_sort_key(slot):
    m = re.match(r"(Junior|Classic|Senior) (Early|Late) (\w+)", slot or "")
    if not m:
        return (99, 9, "zzz")
    phase, half, mon = m.groups()
    phase_ord = {"Junior": 0, "Classic": 1, "Senior": 2}[phase]
    half_ord = {"Early": 0, "Late": 1}[half]
    mon_ord = list(MONTHS.keys()).index(mon) if mon in MONTHS else 12
    return (phase_ord, mon_ord, half_ord)


def extract_races():
    wb = openpyxl.load_workbook(RLUMA, data_only=True)

    # Sheet28 has richer per-race metadata (fans required, url slug, venue detail),
    # but only one row per unique race name (not per career-phase instance).
    ws28 = wb["Sheet28"]
    header28 = [ws28.cell(row=1, column=c).value for c in range(1, 18)]
    idx28 = {h: i + 1 for i, h in enumerate(header28) if h}
    extra_by_name = {}
    for r in range(2, ws28.max_row + 1):
        name = ws28.cell(row=r, column=idx28["Race Name"]).value
        if not name:
            continue
        g = lambda col: clean(ws28.cell(row=r, column=idx28[col]).value) if col in idx28 else None
        extra_by_name[name] = {
            "urlSlug": g("URL Slug"),
            "racecourse": g("Racetrack"),
            "course": g("Course"),
            "direction": g("Direction"),
            "participants": g("Participants"),
            "timeOfDay": g("Time of Day"),
            "fansRequired": g("Fans Required"),
            "fansGained": g("Fans Gained (1st)"),
        }

    # Race Data is the canonical per-instance calendar (413 rows, one per
    # actual race run in the 72-slot 3-year career + finales).
    ws = wb["Race Data"]
    out = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        grade = ws.cell(row=r, column=1).value
        dist_type = ws.cell(row=r, column=3).value
        meters_raw = ws.cell(row=r, column=4).value
        date_index = clean(ws.cell(row=r, column=6).value)
        slot = ws.cell(row=r, column=7).value
        surface = ws.cell(row=r, column=8).value
        racecourse = ws.cell(row=r, column=9).value
        base_fans = clean(ws.cell(row=r, column=14).value)

        meters = None
        if isinstance(meters_raw, str):
            m = re.match(r"(\d+)", meters_raw)
            if m:
                meters = int(m.group(1))
        elif isinstance(meters_raw, (int, float)):
            meters = int(meters_raw)

        phase = None
        if slot:
            m = re.match(r"(Junior|Classic|Senior)", slot)
            if m:
                phase = m.group(1)

        extra = extra_by_name.get(name, {})

        out.append({
            "name": name,
            "grade": grade,
            "distType": dist_type,
            "meters": meters,
            "dateIndex": date_index,
            "slot": slot,
            "phase": phase,
            "surface": surface if surface != "Varies" else None,
            "racecourse": racecourse if racecourse != "Varies" else None,
            "baseFans": base_fans,
            "urlSlug": extra.get("urlSlug"),
            "course": extra.get("course"),
            "direction": extra.get("direction"),
            "participants": extra.get("participants"),
            "timeOfDay": extra.get("timeOfDay"),
            "fansRequired": extra.get("fansRequired"),
            "fansGained": extra.get("fansGained") or base_fans,
        })
    wb.close()

    out.sort(key=lambda x: slot_sort_key(x["slot"]))
    return out


def main():
    OUT_DIR.mkdir(exist_ok=True)

    umas = extract_umas()
    (OUT_DIR / "umas.json").write_text(json.dumps(umas, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"umas.json: {len(umas)} entries")

    supports = extract_supports()
    (OUT_DIR / "supports.json").write_text(json.dumps(supports, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"supports.json: {len(supports)} entries")

    races = extract_races()
    (OUT_DIR / "races.json").write_text(json.dumps(races, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"races.json: {len(races)} entries")


if __name__ == "__main__":
    main()
